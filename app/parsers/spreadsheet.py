from __future__ import annotations

import asyncio
from io import BytesIO

from openpyxl import load_workbook

from app.domain.german import normalize_german_text
from app.domain.models import ParsedDocument, ProcessingStatus, SpreadsheetRow


class SpreadsheetParser:
    version = "spreadsheet_parser:v1"
    supported_mime_types = frozenset(
        {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        }
    )

    async def parse(self, content: bytes, *, file_name: str) -> ParsedDocument:
        rows = await asyncio.to_thread(self._extract_rows, content)
        rendered = "\n\n".join(self.render_row(row) for row in rows)
        return ParsedDocument(
            rows=rows,
            original_text=rendered,
            normalized_text=normalize_german_text(rendered),
            status=ProcessingStatus.READY if rows else ProcessingStatus.FAILED,
            parser_version=self.version,
        )

    @staticmethod
    def _extract_rows(content: bytes) -> list[SpreadsheetRow]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        result: list[SpreadsheetRow] = []
        try:
            for sheet in workbook.worksheets:
                values = list(sheet.iter_rows(values_only=True))
                if not values:
                    continue
                headers = [str(value or "").strip() for value in values[0]]
                for row_number, row in enumerate(values[1:], start=2):
                    rendered = [str(value) if value is not None else "" for value in row]
                    if not any(rendered):
                        continue
                    result.append(
                        SpreadsheetRow(
                            sheet_name=sheet.title,
                            row_number=row_number,
                            column_headers=headers,
                            values=rendered,
                        )
                    )
            return result
        finally:
            workbook.close()

    @staticmethod
    def render_row(row: SpreadsheetRow) -> str:
        fields = [
            f"{header or f'Column {index + 1}'}: {value}"
            for index, (header, value) in enumerate(
                zip(row.column_headers, row.values, strict=False)
            )
            if value
        ]
        return f"Sheet: {row.sheet_name}\nRow: {row.row_number}\n" + "\n".join(fields)
